# finanzmanager/application/services/export_service.py
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal

from src.infrastructure.io.csv_writer import write_dicts_to_csv
from src.infrastructure.unit_of_work import UnitOfWork


@dataclass(frozen=True)
class ExportResult:
    dataset: str
    path: str
    rows: int


class ExportService:
    DATASETS = [
        "accounts",
        "employers",
        "pay_rules",
        "categories",
        "income_fixed",
        "income_hourly",
        "expense_recurring",
        "expense_variable",
        "loans",
        "loan_events",
    ]

    SCHEMAS: dict[str, dict] = {
        "accounts": {"fields": ["label", "account_name", "bank_name", "iban", "role_income", "role_debit", "notes"], "needs_period": False},
        "employers": {"fields": ["name", "payout_timing", "default_account_label", "notes"], "needs_period": False},
        "pay_rules": {"fields": ["employer_name", "rule_type", "unit", "value", "notes"], "needs_period": False},
        "categories": {"fields": ["name", "group"], "needs_period": False},
        "income_fixed": {"fields": ["employer_name", "year", "month", "base_amount", "special_amount", "actual_amount", "payout_timing", "account_label", "notes"], "needs_period": True},
        "income_hourly": {
            "fields": [
                "employer_name", "year", "month",
                "hours_normal", "night", "sunday", "holiday", "overtime",
                "special_amount", "actual_amount", "payout_timing", "account_label", "notes"
            ],
            "needs_period": True,
        },
        "expense_recurring": {"fields": ["name", "category_name", "amount", "frequency_months", "due_day", "anchor_month", "status", "account_label", "pay_bucket", "allocation_override", "notes"], "needs_period": False},
        "expense_variable": {"fields": ["name", "category_name", "amount", "year", "month", "status", "account_label", "pay_bucket", "notes"], "needs_period": True},
        "loans": {"fields": ["name", "start_date", "principal_initial", "annual_interest_rate", "regular_payment", "payment_timing", "account_label", "status", "notes"], "needs_period": False},
        "loan_events": {"fields": ["loan_name", "event_date", "year", "month", "event_type", "amount", "new_regular_payment", "new_annual_interest_rate", "notes"], "needs_period": False},
    }

    def __init__(self, uow_factory=UnitOfWork) -> None:
        self._uow_factory = uow_factory

    def schema_fields(self, dataset: str) -> list[str]:
        if dataset not in self.SCHEMAS:
            raise ValueError(f"Unknown dataset: {dataset}")
        return list(self.SCHEMAS[dataset]["fields"])

    def write_csv_template(self, path: str, dataset: str, *, include_examples: bool = False) -> ExportResult:
        fields = self.schema_fields(dataset)
        rows = self._example_rows(dataset) if include_examples else []
        write_dicts_to_csv(path, rows, fieldnames=fields, delimiter=";")
        return ExportResult(dataset=dataset, path=path, rows=len(rows))

    def write_excel_template(self, path: str) -> ExportResult:
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        def _add_sheet(name: str, headers: list[str], rows: list[list]) -> None:
            ws = wb.create_sheet(name)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for row in rows:
                ws.append(row)

        _add_sheet(
            "Konten",
            ["label", "account_name", "bank_name", "iban", "role_income", "role_debit", "notes"],
            [["GIRO", "Girokonto", "Beispielbank", "DE00123456781234567890", "1", "1", "Hauptkonto"]],
        )
        _add_sheet(
            "Infos",
            ["employer", "payout_timing", "default_account", "rule_type", "unit", "value", "notes"],
            [
                ["Firma Beispiel", "MID", "GIRO", "", "", "", ""],
                ["Firma Beispiel", "", "", "HOURLY_WAGE", "EUR_PER_HOUR", "15.00", "Stundenlohn"],
            ],
        )
        _add_sheet(
            "Einkommen_Fest",
            ["Employer", "Jahr", "Monat", "Grundbetrag", "Sonder", "IST", "Auszahlung", "Konto", "Notiz"],
            [["Firma Beispiel", "2026", "1", "3000.00", "0.00", "0.00", "MID", "GIRO", ""]],
        )
        _add_sheet(
            "Einkommen_Stunden",
            ["Employer", "Jahr", "Monat", "Stunden", "Nacht", "Sonntag", "Feiertag", "Überstunden", "Sonder", "IST", "Auszahlung", "Konto", "Notiz"],
            [["Firma Beispiel", "2026", "1", "160", "10", "0", "0", "5", "0.00", "0.00", "MID", "GIRO", ""]],
        )
        _add_sheet(
            "Abo & Verträge",
            ["Name", "Kategorie", "Betrag", "Intervall", "Tag", "Startmonat", "Status", "Konto", "Zahlungszeitpunkt", "ModusOverride", "Notiz"],
            [["Netflix", "Abos", "12.99", "1", "5", "", "ACTIVE", "GIRO", "NONE", "", ""]],
        )
        _add_sheet(
            "Sonder_Ausgaben",
            ["Name", "Kategorie", "Betrag", "Jahr", "Monat", "Status", "Konto", "Zahlungszeitpunkt", "Notiz"],
            [["Lebensmittel", "Haushalt", "250.00", "2026", "1", "OPEN", "GIRO", "NONE", ""]],
        )
        _add_sheet(
            "Kredit",
            ["Kredit", "Startdatum", "Startbetrag", "Rate", "Konto", "Status", "Auszahlung", "Jahr", "Monat", "Zahlung", "Extra"],
            [
                ["Auto Kredit", "2026-01-01", "15000.00", "250.00", "GIRO", "ACTIVE", "MID", "", "", "", ""],
                ["Auto Kredit", "", "", "", "", "", "", "2026", "1", "250.00", ""],
            ],
        )

        wb.save(path)
        return ExportResult(dataset="excel_template", path=path, rows=7)

    def export_csv(self, path: str, dataset: str, period=None) -> ExportResult:
        if dataset not in self.DATASETS:
            raise ValueError(f"Unknown dataset: {dataset}")

        schema = self.SCHEMAS.get(dataset)
        if schema and schema.get("needs_period") and period is None:
            raise ValueError(f"period required for export '{dataset}'")

        with self._uow_factory() as uow:
            if dataset == "accounts":
                rows = self._export_accounts(uow)
            elif dataset == "employers":
                rows = self._export_employers(uow)
            elif dataset == "pay_rules":
                rows = self._export_pay_rules(uow)
            elif dataset == "categories":
                rows = self._export_categories(uow)
            elif dataset == "income_fixed":
                rows = self._export_income_fixed(uow, period.year, period.month)
            elif dataset == "income_hourly":
                rows = self._export_income_hourly(uow, period.year, period.month)
            elif dataset == "expense_recurring":
                rows = self._export_expense_recurring(uow)
            elif dataset == "expense_variable":
                rows = self._export_expense_variable(uow, period.year, period.month)
            elif dataset == "loans":
                rows = self._export_loans(uow)
            elif dataset == "loan_events":
                rows = self._export_loan_events(uow)
            else:
                raise ValueError(f"Unhandled dataset: {dataset}")

        fields = self.schema_fields(dataset)
        write_dicts_to_csv(path, rows, fieldnames=fields, delimiter=";")
        return ExportResult(dataset=dataset, path=path, rows=len(rows))

    # -------- example rows (Vorschlag b) --------
    def _example_rows(self, dataset: str) -> list[dict]:
        # Intentionally minimal + valid-looking.
        if dataset == "accounts":
            return [{
                "label": "GIRO",
                "account_name": "Girokonto",
                "bank_name": "Beispielbank",
                "iban": "DE00123456781234567890",
                "role_income": "1",
                "role_debit": "1",
                "notes": "Hauptkonto",
            }]
        if dataset == "employers":
            return [{
                "name": "Firma Beispiel",
                "payout_timing": "MID",
                "default_account_label": "GIRO",
                "notes": "",
            }]
        if dataset == "pay_rules":
            return [
                {"employer_name": "Firma Beispiel", "rule_type": "Stundenlohn", "unit": "EUR_PER_HOUR", "value": "15.00", "notes": "BW"},
                {"employer_name": "Firma Beispiel", "rule_type": "Nachtzuschlag", "unit": "MULTIPLIER", "value": "1.25", "notes": ""},
            ]
        if dataset == "categories":
            return [{"name": "Miete", "group": "FIX"}]
        if dataset == "income_fixed":
            return [{
                "employer_name": "Firma Beispiel",
                "year": "2026",
                "month": "1",
                "base_amount": "3000.00",
                "special_amount": "0.00",
                "actual_amount": "0.00",
                "payout_timing": "MID",
                "account_label": "GIRO",
                "notes": "",
            }]
        if dataset == "income_hourly":
            return [{
                "employer_name": "Firma Beispiel",
                "year": "2026",
                "month": "1",
                "hours_normal": "160",
                "night": "10",
                "sunday": "0",
                "holiday": "0",
                "overtime": "5",
                "special_amount": "0.00",
                "actual_amount": "0.00",
                "payout_timing": "MID",
                "account_label": "GIRO",
                "notes": "",
            }]
        if dataset == "expense_recurring":
            return [{
                "name": "Netflix",
                "category_name": "Abos",
                "amount": "12.99",
                "frequency_months": "1",
                "due_day": "5",
                "anchor_month": "",
                "status": "ACTIVE",
                "account_label": "GIRO",
                "pay_bucket": "NONE",
                "allocation_override": "",
                "notes": "",
            }]
        if dataset == "expense_variable":
            return [{
                "name": "Lebensmittel",
                "category_name": "Haushalt",
                "amount": "250.00",
                "year": "2026",
                "month": "1",
                "status": "OPEN",
                "account_label": "GIRO",
                "pay_bucket": "NONE",
                "notes": "",
            }]
        if dataset == "loans":
            return [{
                "name": "Auto Kredit",
                "start_date": "2026-01-01",
                "principal_initial": "15000.00",
                "annual_interest_rate": "0.0000",
                "regular_payment": "250.00",
                "payment_timing": "MID",
                "account_label": "GIRO",
                "status": "ACTIVE",
                "notes": "",
            }]
        if dataset == "loan_events":
            return [{
                "loan_name": "Auto Kredit",
                "event_date": "2026-01-15",
                "year": "2026",
                "month": "1",
                "event_type": "PAYMENT",
                "amount": "250.00",
                "new_regular_payment": "",
                "new_annual_interest_rate": "",
                "notes": "",
            }]
        return []

    # -------- exports --------
    def _export_accounts(self, uow) -> list[dict]:
        out = []
        for a in uow.accounts.list_all():
            out.append({
                "label": a.label,
                "account_name": a.account_name,
                "bank_name": a.bank_name or "",
                "iban": a.iban or "",
                "role_income": "1" if a.role_income else "0",
                "role_debit": "1" if a.role_debit else "0",
                "notes": a.notes or "",
            })
        return out

    def _export_employers(self, uow) -> list[dict]:
        acc_by_id = {a.id: a.label for a in uow.accounts.list_all()}
        out = []
        for e in uow.employers.list_all():
            out.append({
                "name": e.name,
                "payout_timing": getattr(e.payout_timing, "value", str(e.payout_timing)),
                "default_account_label": acc_by_id.get(e.default_account_id, ""),
                "notes": e.notes or "",
            })
        return out

    def _export_pay_rules(self, uow) -> list[dict]:
        emp_by_id = {e.id: e.name for e in uow.employers.list_all()}
        out = []
        for r in uow.pay_rules.list_all():
            out.append({
                "employer_name": emp_by_id.get(r.employer_id, str(r.employer_id)),
                "rule_type": getattr(r.rule_type, "value", str(r.rule_type)),
                "unit": getattr(r.unit, "value", str(r.unit)),
                "value": str(r.value),
                "notes": getattr(r, "notes", None) or "",
            })
        return out

    def _export_categories(self, uow) -> list[dict]:
        out = []
        for c in uow.categories.list_all():
            out.append({"name": c.name, "group": getattr(c.group, "value", str(c.group))})
        return out

    def _export_income_fixed(self, uow, year: int, month: int) -> list[dict]:
        emp_by_id = {e.id: e.name for e in uow.employers.list_all()}
        acc_by_id = {a.id: a.label for a in uow.accounts.list_all()}
        out = []
        for r in uow.income_fixed.list_for_period(year, month):
            out.append({
                "employer_name": emp_by_id.get(r.employer_id, str(r.employer_id)),
                "year": r.year,
                "month": r.month,
                "base_amount": str(r.base_amount),
                "special_amount": str(r.special_amount),
                "actual_amount": str(r.actual_amount),
                "payout_timing": getattr(r.payout_timing, "value", str(r.payout_timing)),
                "account_label": acc_by_id.get(r.account_id, ""),
                "notes": r.notes or "",
            })
        return out

    def _export_income_hourly(self, uow, year: int, month: int) -> list[dict]:
        emp_by_id = {e.id: e.name for e in uow.employers.list_all()}
        acc_by_id = {a.id: a.label for a in uow.accounts.list_all()}
        out = []
        for r in uow.income_hourly.list_for_period(year, month):
            # BW/BY legacy -> neutral export
            hours_normal = (r.hours_normal or Decimal("0")) + (r.hours_bw or Decimal("0")) + (r.hours_by or Decimal("0"))
            night = (r.night or Decimal("0")) + (r.night_bw or Decimal("0")) + (r.night_by or Decimal("0"))
            sunday = (r.sunday or Decimal("0")) + (r.sunday_bw or Decimal("0")) + (r.sunday_by or Decimal("0"))
            out.append({
                "employer_name": emp_by_id.get(r.employer_id, str(r.employer_id)),
                "year": r.year,
                "month": r.month,
                "hours_normal": str(hours_normal),
                "night": str(night),
                "sunday": str(sunday),
                "holiday": str(r.holiday or Decimal("0")),
                "overtime": str(r.overtime or Decimal("0")),
                "special_amount": str(r.special_amount),
                "actual_amount": str(r.actual_amount),
                "payout_timing": getattr(r.payout_timing, "value", str(r.payout_timing)),
                "account_label": acc_by_id.get(r.account_id, ""),
                "notes": r.notes or "",
            })
        return out

    def _export_expense_recurring(self, uow) -> list[dict]:
        cat_by_id = {c.id: c.name for c in uow.categories.list_all()}
        acc_by_id = {a.id: a.label for a in uow.accounts.list_all()}
        out = []
        for r in uow.expense_recurring.list_all():
            out.append({
                "name": r.name,
                "category_name": cat_by_id.get(r.category_id, str(r.category_id)),
                "amount": str(r.amount),
                "frequency_months": r.frequency_months,
                "due_day": r.due_day,
                "anchor_month": r.anchor_month or "",
                "status": getattr(r.status, "value", str(r.status)),
                "account_label": acc_by_id.get(r.account_id, ""),
                "pay_bucket": getattr(r.pay_bucket, "value", str(r.pay_bucket)),
                "allocation_override": (getattr(r.allocation_override, "value", str(r.allocation_override)) if r.allocation_override else ""),
                "notes": r.notes or "",
            })
        return out

    def _export_expense_variable(self, uow, year: int, month: int) -> list[dict]:
        cat_by_id = {c.id: c.name for c in uow.categories.list_all()}
        acc_by_id = {a.id: a.label for a in uow.accounts.list_all()}
        out = []
        for r in uow.expense_variable.list_for_period(year, month):
            out.append({
                "name": r.name,
                "category_name": cat_by_id.get(r.category_id, str(r.category_id)),
                "amount": str(r.amount),
                "year": r.year,
                "month": r.month,
                "status": getattr(r.status, "value", str(r.status)),
                "account_label": acc_by_id.get(r.account_id, ""),
                "pay_bucket": getattr(r.pay_bucket, "value", str(r.pay_bucket)),
                "notes": r.notes or "",
            })
        return out

    def _export_loans(self, uow) -> list[dict]:
        acc_by_id = {a.id: a.label for a in uow.accounts.list_all()}
        out = []
        for l in uow.loans.list_all():
            out.append({
                "name": l.name,
                "start_date": str(l.start_date),
                "principal_initial": str(l.principal_initial),
                "annual_interest_rate": str(l.annual_interest_rate),
                "regular_payment": str(l.regular_payment),
                "payment_timing": getattr(l.payment_timing, "value", str(l.payment_timing)),
                "account_label": acc_by_id.get(l.account_id, ""),
                "status": getattr(l.status, "value", str(l.status)),
                "notes": l.notes or "",
            })
        return out

    def _export_loan_events(self, uow) -> list[dict]:
        loan_by_id = {l.id: l.name for l in uow.loans.list_all()}
        out = []
        for e in uow.loan_events.list_all():
            out.append({
                "loan_name": loan_by_id.get(e.loan_id, str(e.loan_id)),
                "event_date": str(e.event_date),
                "year": e.year,
                "month": e.month,
                "event_type": getattr(e.event_type, "value", str(e.event_type)),
                "amount": (str(e.amount) if e.amount is not None else ""),
                "new_regular_payment": (str(e.new_regular_payment) if e.new_regular_payment is not None else ""),
                "new_annual_interest_rate": (str(e.new_annual_interest_rate) if e.new_annual_interest_rate is not None else ""),
                "notes": e.notes or "",
            })
        return out
