# infrastructure/io/excel_reader.py
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ExcelSheet:
    name: str
    rows: list[dict[str, Any]]


def list_sheets(path: str | Path) -> list[str]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))

    wb = load_workbook(filename=str(p), data_only=True, keep_vba=p.suffix.lower() == ".xlsm")
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_sheet_dicts(
    path: str | Path,
    sheet_name: str,
    header_row: int = 1,
    first_data_row: int | None = None,
    max_rows: int | None = None,
) -> list[dict[str, Any]]:
    """
    Reads an Excel sheet into list[dict] using header_row as keys.

    - header_row is 1-based (Excel)
    - first_data_row defaults to header_row + 1
    - Empty header cells are ignored.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))

    wb = load_workbook(filename=str(p), data_only=True, keep_vba=p.suffix.lower() == ".xlsm")
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]

        hdr_row = header_row
        data_row_start = first_data_row or (hdr_row + 1)

        headers: list[str] = []
        for cell in ws[hdr_row]:
            h = "" if cell.value is None else str(cell.value).strip()
            headers.append(h)

        rows: list[dict[str, Any]] = []
        count = 0
        for r_idx in range(data_row_start, ws.max_row + 1):
            if max_rows is not None and count >= max_rows:
                break

            values = [ws.cell(row=r_idx, column=c_idx + 1).value for c_idx in range(len(headers))]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
                continue

            row: dict[str, Any] = {}
            for h, v in zip(headers, values):
                if not h:
                    continue
                row[h] = v
            rows.append(row)
            count += 1

        return rows
    finally:
        wb.close()


def read_excel_template(
    path: str | Path,
    sheet_names: list[str],
    header_row: int = 1,
) -> list[ExcelSheet]:
    """
    Convenience: load multiple sheets.
    """
    out: list[ExcelSheet] = []
    for name in sheet_names:
        out.append(ExcelSheet(name=name, rows=read_sheet_dicts(path, name, header_row=header_row)))
    return out
